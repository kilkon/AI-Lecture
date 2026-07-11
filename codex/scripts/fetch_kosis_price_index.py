from __future__ import annotations

import io
import json
import re
import sys
import time
import zipfile
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, build_opener

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
KEY_PATH = ROOT / "kosis_api_key.txt"
OUT = ROOT / "data" / "kosis_price_index.xlsx"

KOSIS_DATA_URL = "https://kosis.kr/openapi/Param/statisticsParameterData.do"
KOSIS_SEARCH_URL = "https://kosis.kr/search/searchStatDBAjax.do"
KOSIS_METADATA_ZIP = "https://kosis.kr/downXLS/ZTITLE.zip"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
}


def read_api_key() -> str:
    if not KEY_PATH.exists():
        raise FileNotFoundError(f"KOSIS API key file not found: {KEY_PATH}")
    key = KEY_PATH.read_text(encoding="utf-8").strip().strip('"').strip("'")
    if not key:
        raise ValueError(f"KOSIS API key file is empty: {KEY_PATH}")
    return key


def session():
    return build_opener()


def request_json(sess, method: str, url: str, **kwargs: Any) -> Any:
    last_error: Exception | None = None
    for attempt in range(5):
        try:
            params = kwargs.get("params")
            data = kwargs.get("data")
            target_url = url
            body = None
            headers = dict(HEADERS)
            if method.upper() == "GET" and params:
                target_url = f"{url}?{urlencode(params)}"
            if method.upper() == "POST" and data:
                body = urlencode(data).encode("utf-8")
                headers["Content-Type"] = "application/x-www-form-urlencoded"
            request = Request(target_url, data=body, headers=headers, method=method.upper())
            with sess.open(request, timeout=60) as response:
                raw = response.read().decode("utf-8-sig")
            return json.loads(raw)
        except (HTTPError, URLError, TimeoutError, ValueError) as exc:
            last_error = exc
            if attempt == 4:
                break
            time.sleep(min(8.0, 0.4 * (2**attempt)))
    raise RuntimeError(f"KOSIS request failed: {last_error}")


def table_id_from_row(row: dict[str, Any]) -> str | None:
    for key, value in row.items():
        if "TBL" in key.upper() and "ID" in key.upper() and value:
            text = str(value).strip()
            if text:
                return text
    text = " ".join(str(v) for v in row.values())
    match = re.search(r"\bDT_[A-Z0-9_]+\b", text, re.IGNORECASE)
    return match.group(0).upper() if match else None


def title_from_row(row: dict[str, Any]) -> str:
    parts: list[str] = []
    for value in row.values():
        if isinstance(value, str) and "소비자물가지수" in value:
            parts.append(value.strip())
    return " | ".join(parts[:3])


def search_table_ids(sess) -> list[dict[str, str]]:
    candidates: dict[str, dict[str, str]] = {}
    for gbn in ("L", "E", "I", "B"):
        try:
            data = request_json(
                sess,
                "POST",
                KOSIS_SEARCH_URL,
                data={"query": "소비자물가지수", "gbn": gbn},
            )
        except RuntimeError:
            continue
        for row in data.get("resultList", []) if isinstance(data, dict) else []:
            tbl = table_id_from_row(row)
            if tbl:
                candidates.setdefault(
                    tbl,
                    {
                        "tbl_id": tbl,
                        "org_id": str(row.get("ORG_ID", "")).strip(),
                        "title": title_from_row(row),
                        "source": "search",
                    },
                )
    return list(candidates.values())


def metadata_table_ids(sess) -> list[dict[str, str]]:
    request = Request(KOSIS_METADATA_ZIP, headers=HEADERS, method="GET")
    with sess.open(request, timeout=90) as response:
        content = response.read()
    out: dict[str, dict[str, str]] = {}
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        excel_name = next(
            name for name in zf.namelist() if name.lower().endswith((".xlsx", ".xls"))
        )
        with zf.open(excel_name) as ef:
            meta = pd.read_excel(ef, header=3)
    for _, row in meta.iterrows():
        values = [str(v).strip() for v in row.tolist() if pd.notna(v)]
        text = " | ".join(values)
        if "소비자물가지수" not in text:
            continue
        match = re.search(r"\bDT_[A-Z0-9_]+\b", text, re.IGNORECASE)
        if not match:
            continue
        tbl = match.group(0).upper()
        out.setdefault(
            tbl,
            {
                "tbl_id": tbl,
                "org_id": "",
                "title": text[:240],
                "source": "metadata",
            },
        )
    return list(out.values())


def score_candidate(candidate: dict[str, str]) -> tuple[int, str]:
    text = (candidate.get("title") or "") + " " + candidate.get("tbl_id", "")
    score = 0
    if "소비자물가지수" in text:
        score += 20
    if "품목" in text or "지출목적" in text:
        score += 4
    if "시도" in text or "지역" in text:
        score -= 4
    if "전년" in text or "등락" in text or "증감" in text:
        score -= 6
    if candidate.get("org_id"):
        score += 2
    return (-score, candidate.get("tbl_id", ""))


def lookup_org_id(sess, tbl_id: str) -> str:
    for gbn in ("L", "E", "I", "B"):
        data = request_json(
            sess, "POST", KOSIS_SEARCH_URL, data={"query": tbl_id, "gbn": gbn}
        )
        rows = data.get("resultList", []) if isinstance(data, dict) else []
        for row in rows:
            org = str(row.get("ORG_ID", "")).strip()
            if org:
                return org
    if len(tbl_id) > 6 and tbl_id[3:6].isdigit():
        return str(int(tbl_id[3:6]))
    raise LookupError(f"Could not find orgId for {tbl_id}")


def normalize_payload(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict) and "StatisticalData" in payload:
        data = payload["StatisticalData"]
        if isinstance(data, list):
            return data
    if isinstance(payload, dict) and "err" in payload:
        raise RuntimeError(f"KOSIS error {payload.get('err')}: {payload.get('errMsg')}")
    raise RuntimeError(f"Unexpected KOSIS payload: {type(payload).__name__}")


def fetch_table(sess, api_key: str, candidate: dict[str, str]) -> pd.DataFrame:
    tbl_id = candidate["tbl_id"]
    org_id = candidate.get("org_id") or lookup_org_id(sess, tbl_id)
    base_params = {
        "method": "getList",
        "apiKey": api_key,
        "format": "json",
        "jsonVD": "Y",
        "orgId": org_id,
        "tblId": tbl_id,
        "prdSe": "M",
        "newEstPrdCnt": "120",
        "itmId": "ALL",
    }
    last_error: Exception | None = None
    for obj_count in range(1, 9):
        params = dict(base_params)
        for idx in range(1, obj_count + 1):
            params[f"objL{idx}"] = "ALL"
        try:
            payload = request_json(sess, "GET", KOSIS_DATA_URL, params=params)
            rows = normalize_payload(payload)
            df = pd.DataFrame(rows)
            if not df.empty and {"PRD_DE", "DT"}.issubset(df.columns):
                candidate["org_id"] = org_id
                return df
        except Exception as exc:
            last_error = exc
            message = str(exc)
            if "KOSIS error 20" in message:
                continue
            if "KOSIS error 21" in message and obj_count > 1:
                break
    raise RuntimeError(f"{tbl_id} fetch failed: {last_error}")


def category_columns(df: pd.DataFrame) -> list[str]:
    cols = [c for c in df.columns if c.endswith("_NM") or re.fullmatch(r"C\d+_NM", c)]
    return [c for c in cols if c not in {"TBL_NM"}]


def build_trend(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    work["price_index"] = pd.to_numeric(work["DT"], errors="coerce")
    work["period"] = work["PRD_DE"].astype(str)
    text_cols = category_columns(work)
    if text_cols:
        joined = work[text_cols].astype(str).agg(" | ".join, axis=1)
        include = joined.str.contains("총지수|소비자물가지수", regex=True, na=False)
        exclude = joined.str.contains("전년|전월|등락|증감|기여도", regex=True, na=False)
        filtered = work[include & ~exclude].copy()
        if not filtered.empty:
            work = filtered
    trend = (
        work.dropna(subset=["price_index", "period"])
        .groupby("period", as_index=False)["price_index"]
        .mean()
        .sort_values("period")
    )
    if trend.empty:
        raise RuntimeError("No numeric price index rows were found.")
    trend["year"] = trend["period"].str.slice(0, 4)
    trend["month"] = trend["period"].str.slice(4, 6)
    trend["period_label"] = trend["year"] + "-" + trend["month"]
    trend["period_date"] = pd.to_datetime(trend["period_label"] + "-01", errors="coerce")
    return trend[["period", "period_label", "period_date", "price_index"]]


def write_workbook(raw: pd.DataFrame, trend: pd.DataFrame, candidate: dict[str, str]) -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        info = pd.DataFrame(
            [
                ["통계표 ID", candidate.get("tbl_id", "")],
                ["기관 ID", candidate.get("org_id", "")],
                ["자료 출처", "KOSIS OpenAPI"],
                ["기간", f"{trend['period_label'].min()} ~ {trend['period_label'].max()}"],
                ["행 수", len(trend)],
            ],
            columns=["항목", "값"],
        )
        info.to_excel(writer, sheet_name="summary", index=False)
        trend.to_excel(writer, sheet_name="price_trend", index=False)
        raw.to_excel(writer, sheet_name="raw_kosis", index=False)

    wb = load_workbook(OUT)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F766E")
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells[:80])
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                max(max_len + 2, 10), 34
            )

    ws = wb["price_trend"]
    ws["C1"] = "date_for_axis"
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).number_format = "yyyy-mm"
    chart = LineChart()
    chart.title = "KOSIS 월별 물가지수 추세"
    chart.y_axis.title = "물가지수"
    chart.x_axis = DateAxis()
    chart.x_axis.title = "날짜"
    chart.x_axis.number_format = "yyyy-mm"
    chart.x_axis.majorTimeUnit = "months"
    chart.x_axis.tickLblPos = "low"
    data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 18
    ws.add_chart(chart, "E2")
    wb.save(OUT)


def main() -> int:
    api_key = read_api_key()
    sess = session()
    candidates = search_table_ids(sess)
    try:
        candidates.extend(metadata_table_ids(sess))
    except Exception:
        pass
    unique = {c["tbl_id"]: c for c in candidates if c.get("tbl_id")}
    ordered = sorted(unique.values(), key=score_candidate)
    if not ordered:
        raise RuntimeError("No KOSIS consumer price index table candidates found.")

    errors: list[str] = []
    for candidate in ordered[:20]:
        try:
            raw = fetch_table(sess, api_key, candidate)
            trend = build_trend(raw)
            write_workbook(raw, trend, candidate)
            print(f"created={OUT}")
            print(f"table_id={candidate.get('tbl_id')}")
            print(f"org_id={candidate.get('org_id')}")
            print(f"rows={len(trend)}")
            print(f"period={trend['period_label'].min()}..{trend['period_label'].max()}")
            return 0
        except Exception as exc:
            errors.append(f"{candidate.get('tbl_id')}: {exc}")
    raise RuntimeError("All candidate tables failed:\n" + "\n".join(errors[:10]))


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
